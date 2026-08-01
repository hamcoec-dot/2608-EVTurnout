import os
import re
import csv
import json
from datetime import datetime, timedelta
import openpyxl
import fitz # PyMuPDF

def parse_date(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        val = val.strip()
        # Check if it matches YYYY-MM-DD
        if re.match(r'^\d{4}-\d{2}-\d{2}$', val):
            return val
        # Check if it is a datetime string like 2026-07-17 00:00:00
        try:
            dt = datetime.strptime(val, '%Y-%m-%d %H:%M:%S')
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass
    return str(val)

def dates_match(csv_date, excel_label):
    if csv_date == excel_label:
        return True
    m = re.search(r'(\d+/\d+/\d+)', excel_label)
    if m:
        date_part = m.group(1)
        for fmt in ('%m/%d/%Y', '%m/%d/%y'):
            try:
                dt = datetime.strptime(date_part, fmt)
                if dt.strftime('%Y-%m-%d') == csv_date:
                    return True
            except ValueError:
                pass
    return False

def clean_num(val):
    if val is None or val == 'N/A' or val == '':
        return None
    try:
        return int(float(val))
    except ValueError:
        return str(val)

# Helper to parse a historical PDF report using fitz
def parse_historical_pdf(path):
    try:
        doc = fitz.open(path)
        text = doc[0].get_text()
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        # Regex matching dates: Thru X/Y/Z, X/Y/Z, X/Y/Z-Election Day, etc., and supporting 2020's specific format
        date_pattern = re.compile(
            r'^(Thru\s+\d+/\d+/\d+|\d+/\d+(?:/\d+)?(?:\s*-\s*Election Day)?|\d+/\d+(?:/\d+)?-Election Day|Election Day|\d+-\d+\s+thru\s+\d+-\d+|\d+/\d+/\d+\s*-\s*\d+/\d+/\d+)$', 
            re.IGNORECASE
        )
        
        records = []
        current_date = None
        current_values = []
        
        # Extracted party sums
        republican = 0
        democrat = 0
        general = 0
        grand_total = 0
        
        for i, line in enumerate(lines):
            if date_pattern.match(line):
                if current_date:
                    records.append((current_date, current_values))
                current_date = line
                current_values = []
            elif current_date:
                if line.replace('.', '', 1).isdigit() or line == 'N/A':
                    current_values.append(line)
                else:
                    records.append((current_date, current_values))
                    current_date = None
                    current_values = []
            
        # Detect and parse summary party totals
        grouped_found = False
        for k in range(len(lines) - 7):
            if lines[k] == 'REP' and lines[k+1] == 'DEM' and lines[k+2] == 'GEN' and lines[k+3] == 'GRAND TOTAL':
                republican = clean_num(lines[k+4])
                democrat = clean_num(lines[k+5])
                general = clean_num(lines[k+6])
                grand_total = clean_num(lines[k+7])
                grouped_found = True
                break
                
        if not grouped_found:
            for i, line in enumerate(lines):
                lower_line = line.lower()
                if 'republican primary' in lower_line:
                    val = clean_num(lines[i+1]) if i+1 < len(lines) else 0
                    if isinstance(val, int): republican = val
                elif 'democratic primary' in lower_line:
                    val = clean_num(lines[i+1]) if i+1 < len(lines) else 0
                    if isinstance(val, int): democrat = val
                elif 'general only' in lower_line or 'general election' in lower_line:
                    val = clean_num(lines[i+1]) if i+1 < len(lines) else 0
                    if isinstance(val, int): general = val
                elif 'grand total' in lower_line:
                    val = clean_num(lines[i+1]) if i+1 < len(lines) else 0
                    if isinstance(val, int): grand_total = val
                    
        if current_date:
            records.append((current_date, current_values))
            
        # Compile daily totals
        daily_totals = []
        for date, vals in records:
            if vals:
                row_total = clean_num(vals[-1])
                if isinstance(row_total, int):
                    daily_totals.append(row_total)
                    
        # Compute cumulative totals
        cumulative = []
        curr = 0
        for val in daily_totals:
            curr += val
            cumulative.append(curr)
            
        if grand_total == 0 and cumulative:
            grand_total = cumulative[-1]
            
        return {
            "dailyTotals": daily_totals,
            "cumulativeTotals": cumulative,
            "summary": {
                "republican": republican,
                "democrat": democrat,
                "general": general,
                "grandTotal": grand_total
            }
        }
    except Exception as e:
        print(f"Error parsing historical PDF {path}: {e}")
        return None

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    if base_dir:
        os.chdir(base_dir)

    excel_path = 'Early Voting & Absentee Daily Turnout Report - Full County.xlsx'
    config_path = 'config.json'
    
    # 1. Load config.json if present
    config = {}
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
            
    # Default locations list if Excel is not present
    DEFAULT_LOCATIONS = [
        "By Mail/NH Voters",
        "Brainerd",
        "Collegedale",
        "Election Comm.",
        "Harrison",
        "Hixson",
        "Soddy Daisy"
    ]
    
    locations = DEFAULT_LOCATIONS
    report_title = config.get("electionTitle", "Hamilton TN Early Voting & Absentee Daily Turnout Report")
    election_name = config.get("electionName", "State & Federal Primary and County General")
    election_date = config.get("electionDate", "August 6th, 2026")
    total_registered = config.get("totalRegisteredVoters", 218764)
    disclaimer = "*All totals are unofficial and are subject to verification and certification by the Hamilton County Election Commission."
    
    excel_days = {}
    excel_loaded = False
    
    # Load Excel data if present
    if os.path.exists(excel_path):
        try:
            wb_data = openpyxl.load_workbook(excel_path, data_only=True)
            ws_data = wb_data.active
            
            # Resolve metadata with config.json overrides or Excel
            title_val = ws_data['A1'].value or ""
            title_lines = [line.strip() for line in title_val.split('\n') if line.strip()]
            
            if "electionTitle" not in config and len(title_lines) > 0:
                report_title = title_lines[0]
            if "electionName" not in config and len(title_lines) > 1:
                election_name = title_lines[1]
            if "electionDate" not in config and len(title_lines) > 2:
                election_date = title_lines[2]
            if "totalRegisteredVoters" not in config:
                val = clean_num(ws_data.cell(row=24, column=2).value)
                if val: total_registered = val
                
            excel_disclaimer = ws_data.cell(row=29, column=1).value
            if excel_disclaimer:
                disclaimer = excel_disclaimer.strip()
                
            # Load locations from row 5, columns B to H
            excel_locations = []
            for col in range(2, 9):
                val = ws_data.cell(row=5, column=col).value
                if val:
                    excel_locations.append(val.strip())
            if excel_locations:
                locations = excel_locations
                
            # Parse daily values
            for r in range(6, 22):
                date_raw = ws_data.cell(row=r, column=1).value
                if date_raw is None:
                    continue
                date_str = parse_date(date_raw)
                
                values = {}
                for c_idx, loc in enumerate(locations, 2):
                    val = ws_data.cell(row=r, column=c_idx).value
                    values[loc] = clean_num(val)
                    
                total_val = clean_num(ws_data.cell(row=r, column=9).value)
                
                excel_days[date_str] = {
                    "values": values,
                    "total": total_val
                }
            excel_loaded = True
            print("Successfully loaded historical details from Excel sheet")
        except Exception as e:
            print(f"Warning: Failed to load Excel file: {e}")
            
    # Location mapping from CSV to Excel
    LOCATION_MAP = {
        "AB": "By Mail/NH Voters",
        "BR": "Brainerd",
        "CD": "Collegedale",
        "EC": "Election Comm.",
        "HR": "Harrison",
        "HS": "Hixson",
        "SD": "Soddy Daisy",
        "ABSENTEE/BY MAIL": "By Mail/NH Voters",
        "BRAINERD REC CTR": "Brainerd",
        "COLLEGEDALE": "Collegedale",
        "ELECTION COMMISSION": "Election Comm.",
        "HARRISON": "Harrison",
        "HIXSON": "Hixson",
        "SODDY DAISY": "Soddy Daisy"
    }

    PARTY_MAP = {
        "R": "republican",
        "REP": "republican",
        "REPUBLICAN": "republican",
        "D": "democrat",
        "DEM": "democrat",
        "DEMOCRAT": "democrat",
        "G": "general",
        "GEN": "general",
        "GENERAL": "general"
    }

    # Scan for voter CSV files in current directory and Current_Election_Results
    voter_records = []
    seen_registrations = set()
    csv_directories = ['.', 'Current_Election_Results']

    for folder in csv_directories:
        if not os.path.exists(folder):
            continue
        for filename in os.listdir(folder):
            if filename.endswith('.csv'):
                filepath = os.path.join(folder, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                        reader = csv.DictReader(f)
                        fieldnames = reader.fieldnames or []
                        if 'DateBallotReceived' not in fieldnames:
                            continue
                        
                        for row in reader:
                            reg_num = row.get('RegistrationNum', '').strip()
                            if reg_num and reg_num in seen_registrations:
                                continue
                            if reg_num:
                                seen_registrations.add(reg_num)

                            raw_date = row.get('DateBallotReceived', '').strip()
                            if not raw_date:
                                continue

                            parsed_date_str = None
                            for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y', '%m-%d-%Y'):
                                try:
                                    dt = datetime.strptime(raw_date, fmt)
                                    parsed_date_str = dt.strftime('%Y-%m-%d')
                                    break
                                except ValueError:
                                    pass

                            if not parsed_date_str:
                                continue

                            loc_raw = row.get('EarlyVoterLocation', '').strip()
                            mapped_loc = LOCATION_MAP.get(loc_raw, loc_raw)

                            party_raw = row.get('Party', '').strip().upper()
                            mapped_party = PARTY_MAP.get(party_raw, 'general')

                            age_val = None
                            try:
                                age_val = int(row.get('AGE', '').strip())
                            except ValueError:
                                pass

                            sex_val = row.get('Sex', '').strip().upper()
                            if sex_val not in ('F', 'M'):
                                sex_val = 'Other/Unknown'

                            is_ft_voter = (row.get('FirstTimeVoter', '').strip().upper() == 'YES')

                            voter_records.append({
                                "date": parsed_date_str,
                                "location": mapped_loc,
                                "party": mapped_party,
                                "age": age_val,
                                "sex": sex_val,
                                "isFirstTime": is_ft_voter,
                                "precinct": row.get('PrecinctName', '').strip(),
                                "commission": row.get('Commission', '').strip(),
                                "senate": row.get('Senate', '').strip(),
                                "house": row.get('House', '').strip(),
                                "school": row.get('School', '').strip(),
                                "city": row.get('City', '').strip(),
                                "municipality": row.get('Municipality', '').strip()
                            })
                except Exception as e:
                    print(f"Warning: Failed to parse CSV file {filepath}: {e}")

    # Build csv_data daily breakdown from voter records
    csv_data = {}
    for record in voter_records:
        d = record['date']
        loc = record['location']
        party = record['party']

        if d not in csv_data:
            csv_data[d] = {}
        if loc not in csv_data[d]:
            csv_data[d][loc] = {"total": 0, "democrat": 0, "republican": 0, "general": 0}

        csv_data[d][loc]["total"] += 1
        csv_data[d][loc][party] += 1

    # 3. Generate the daily timeline based on configuration dates
    daily_turnout = []
    early_voting_start = config.get("earlyVotingStartDate", "2026-07-17")
    early_voting_end = config.get("earlyVotingEndDate", "2026-08-01")
    holidays = set(config.get("holidays", []))

    timeline_dates = []

    # Generate timeline
    start_dt = datetime.strptime(early_voting_start, "%Y-%m-%d")
    end_dt = datetime.strptime(early_voting_end, "%Y-%m-%d")

    # Day 0: Pre-voting Mail-in block
    pre_ev_date = start_dt - timedelta(days=1)
    pre_ev_label = f"Thru {pre_ev_date.strftime('%m/%d/%Y').replace('/0', '/')}"
    timeline_dates.append(pre_ev_label)

    curr_dt = start_dt
    while curr_dt <= end_dt:
        if curr_dt.weekday() != 6:
            date_str = curr_dt.strftime("%Y-%m-%d")
            if date_str not in holidays:
                timeline_dates.append(date_str)
        curr_dt += timedelta(days=1)

    # Election Day
    election_day_dt = end_dt + timedelta(days=1)
    short_date = f"{election_day_dt.month}/{election_day_dt.day}/{str(election_day_dt.year)[2:]}"
    timeline_dates.append(f"{short_date}-Election Day")

    # Build dailyTurnout objects
    for d_str in timeline_dates:
        is_mail_only = True
        values = {}

        if d_str in excel_days:
            values = excel_days[d_str]["values"]
            total_val = excel_days[d_str]["total"]
            for c_idx, loc in enumerate(locations, 2):
                val = values[loc]
                if val is not None and val != 'N/A' and c_idx > 2:
                    is_mail_only = False
        else:
            total_val = 0
            if "Thru" in d_str or "Election Day" in d_str:
                is_mail_only = True
                values = {loc: (0 if c_idx == 2 else None) for c_idx, loc in enumerate(locations, 2)}
            else:
                is_mail_only = False
                values = {loc: 0 for loc in locations}

        # Aggregate CSV records for this timeline date
        party_breakdown = {}
        for csv_date, breakdown in csv_data.items():
            # Check if csv_date matches pre-EV, exact EV date, or election day
            is_match = False
            if "Thru" in d_str:
                if csv_date < early_voting_start:
                    is_match = True
            elif "Election Day" in d_str:
                if csv_date > early_voting_end:
                    is_match = True
            else:
                if csv_date == d_str:
                    is_match = True

            if is_match:
                for loc, counts in breakdown.items():
                    if loc not in party_breakdown:
                        party_breakdown[loc] = {"total": 0, "democrat": 0, "republican": 0, "general": 0}
                    party_breakdown[loc]["total"] += counts["total"]
                    party_breakdown[loc]["democrat"] += counts["democrat"]
                    party_breakdown[loc]["republican"] += counts["republican"]
                    party_breakdown[loc]["general"] += counts["general"]

        if party_breakdown:
            total_val = 0
            is_mail_only = True
            for c_idx, loc in enumerate(locations, 2):
                if loc in party_breakdown:
                    csv_loc_total = party_breakdown[loc]["total"]
                    values[loc] = csv_loc_total
                    if csv_loc_total is not None:
                        total_val += csv_loc_total
                        if csv_loc_total > 0 and c_idx > 2:
                            is_mail_only = False
                else:
                    if values.get(loc) is None:
                        values[loc] = None

        daily_turnout.append({
            "date": d_str,
            "isMailOnly": is_mail_only,
            "values": values,
            "total": total_val,
            "partyBreakdown": party_breakdown if party_breakdown else None
        })

    # Re-calculate totals row and columns
    totals = {loc: 0 for loc in locations}
    totals["total"] = 0

    for entry in daily_turnout:
        is_election_day_row = "Election Day" in entry["date"]
        for c_idx, loc in enumerate(locations, 2):
            val = entry["values"][loc]
            if isinstance(val, int):
                if c_idx == 2 or not is_election_day_row:
                    totals[loc] += val
        if isinstance(entry["total"], int):
            totals["total"] += entry["total"]

    early_voting_sum = sum(totals[loc] for c_idx, loc in enumerate(locations, 2) if c_idx > 2)
    absentee_sum = totals[locations[0]]
    grand_total_calc = totals["total"]

    # Recalculate party totals from all daily CSVs
    republican_sum = 0
    democrat_sum = 0
    general_sum = 0
    has_any_breakdown = False

    for entry in daily_turnout:
        pb = entry["partyBreakdown"]
        if pb:
            has_any_breakdown = True
            for loc in pb:
                republican_sum += pb[loc].get("republican", 0)
                democrat_sum += pb[loc].get("democrat", 0)
                general_sum += pb[loc].get("general", 0)

    if has_any_breakdown:
        republican_total = republican_sum
        democrat_total = democrat_sum
        general_total = general_sum
        party_total_sum = republican_total + democrat_total + general_total
    else:
        republican_total = 0
        democrat_total = 0
        general_total = 0
        party_total_sum = 0

    turnout_percent_calc = 0.0
    if total_registered and total_registered > 0:
        turnout_percent_calc = grand_total_calc / total_registered

    # Helper for demographic grouping
    def init_party_dict():
        return {"total": 0, "republican": 0, "democrat": 0, "general": 0}

    def add_voter(d_dict, key, party):
        if key not in d_dict:
            d_dict[key] = init_party_dict()
        d_dict[key]["total"] += 1
        if party in d_dict[key]:
            d_dict[key][party] += 1
        else:
            d_dict[key]["general"] += 1

    def get_age_bracket(age):
        if age is None:
            return "Unknown"
        if age < 18:
            return "<18"
        elif age <= 29:
            return "18-29"
        elif age <= 49:
            return "30-49"
        elif age <= 64:
            return "50-64"
        else:
            return "65+"

    def add_nested_voter(d_dict, key1, key2, party):
        if key1 not in d_dict:
            d_dict[key1] = {}
        add_voter(d_dict[key1], key2, party)

    age_groups = {}
    sex_dist = {}
    commission_dist = {}
    senate_dist = {}
    house_dist = {}
    school_dist = {}
    city_dist = {}
    precinct_dist = {}

    precincts_by_location = {}
    precincts_by_age_group = {}
    precincts_by_district = {
        "commission": {},
        "senate": {},
        "house": {},
        "school": {},
        "city": {}
    }

    for v in voter_records:
        p = v["party"]
        age_b = get_age_bracket(v["age"])
        add_voter(age_groups, age_b, p)
        add_voter(sex_dist, v["sex"], p)
        if v["commission"]: add_voter(commission_dist, v["commission"], p)
        if v["senate"]: add_voter(senate_dist, v["senate"], p)
        if v["house"]: add_voter(house_dist, v["house"], p)
        if v["school"]: add_voter(school_dist, v["school"], p)
        if v["city"]: add_voter(city_dist, v["city"], p)
        if v["precinct"]:
            prec = v["precinct"]
            add_voter(precinct_dist, prec, p)

            if v["location"]:
                add_nested_voter(precincts_by_location, v["location"], prec, p)
            add_nested_voter(precincts_by_age_group, age_b, prec, p)

            if v["commission"]: add_nested_voter(precincts_by_district["commission"], v["commission"], prec, p)
            if v["senate"]: add_nested_voter(precincts_by_district["senate"], v["senate"], prec, p)
            if v["house"]: add_nested_voter(precincts_by_district["house"], v["house"], prec, p)
            if v["school"]: add_nested_voter(precincts_by_district["school"], v["school"], prec, p)
            if v["city"]: add_nested_voter(precincts_by_district["city"], v["city"], prec, p)

    demographics = {
        "ageGroups": age_groups,
        "sex": sex_dist,
        "districts": {
            "commission": commission_dist,
            "senate": senate_dist,
            "house": house_dist,
            "school": school_dist,
            "city": city_dist
        },
        "precincts": precinct_dist,
        "precinctsByLocation": precincts_by_location,
        "precinctsByAgeGroup": precincts_by_age_group,
        "precinctsByDistrict": precincts_by_district
    }

    # Aggregate First-Time Hamilton County Voters statistics
    ft_records = [v for v in voter_records if v.get("isFirstTime")]
    ft_total = len(ft_records)
    ft_pct_of_turnout = (ft_total / grand_total_calc * 100) if grand_total_calc > 0 else 0.0

    ft_party_breakdown = {"republican": 0, "democrat": 0, "general": 0}
    ft_age_groups = {}
    ft_sex_dist = {}
    ft_location_dist = {}
    ft_daily_trend = {}
    ft_commission_dist = {}
    ft_senate_dist = {}
    ft_house_dist = {}
    ft_school_dist = {}
    ft_city_dist = {}
    ft_precinct_dist = {}

    for v in ft_records:
        p = v["party"]
        ft_party_breakdown[p] = ft_party_breakdown.get(p, 0) + 1
        add_voter(ft_age_groups, get_age_bracket(v["age"]), p)
        add_voter(ft_sex_dist, v["sex"], p)
        
        loc = v["location"]
        ft_location_dist[loc] = ft_location_dist.get(loc, 0) + 1
        
        d = v["date"]
        ft_daily_trend[d] = ft_daily_trend.get(d, 0) + 1
        
        if v["commission"]: add_voter(ft_commission_dist, v["commission"], p)
        if v["senate"]: add_voter(ft_senate_dist, v["senate"], p)
        if v["house"]: add_voter(ft_house_dist, v["house"], p)
        if v["school"]: add_voter(ft_school_dist, v["school"], p)
        if v["city"]: add_voter(ft_city_dist, v["city"], p)
        if v["precinct"]: add_voter(ft_precinct_dist, v["precinct"], p)

    first_time_voters = {
        "total": ft_total,
        "turnoutPercent": round(ft_pct_of_turnout, 2),
        "partyBreakdown": ft_party_breakdown,
        "ageGroups": ft_age_groups,
        "sex": ft_sex_dist,
        "locations": ft_location_dist,
        "dailyTrend": ft_daily_trend,
        "districts": {
            "commission": ft_commission_dist,
            "senate": ft_senate_dist,
            "house": ft_house_dist,
            "school": ft_school_dist,
            "city": ft_city_dist
        },
        "precincts": ft_precinct_dist
    }
        
    # 4. Parse historical election results PDFs
    historical = {}
    past_elections_conf = config.get("pastElections", {})
    past_results_dir = 'Past_Election_Results'
    
    if os.path.exists(past_results_dir):
        for filename in os.listdir(past_results_dir):
            if filename.endswith('.pdf'):
                year_match = re.search(r'\b(20\d{2})\b', filename)
                if not year_match:
                    year_prefix = re.match(r'^(\d{2})08', filename)
                    if year_prefix:
                        year = "20" + year_prefix.group(1)
                    else:
                        continue
                else:
                    year = year_match.group(1)
                    
                path = os.path.join(past_results_dir, filename)
                res = parse_historical_pdf(path)
                
                if res:
                    year_conf = past_elections_conf.get(year, {})
                    reg_voters = year_conf.get("totalRegisteredVoters")
                    
                    if not reg_voters:
                        if year == "2024": reg_voters = 231431
                        elif year == "2022": reg_voters = 235853
                        elif year == "2020": reg_voters = 222004
                        elif year == "2018": reg_voters = 197052
                        else: reg_voters = 200000
                        
                    res["summary"]["registered"] = reg_voters
                    if reg_voters > 0:
                        res["summary"]["turnoutPercent"] = res["summary"]["grandTotal"] / reg_voters
                    else:
                        res["summary"]["turnoutPercent"] = 0.0
                        
                    if res["summary"]["republican"] == 0:
                        tot = res["summary"]["grandTotal"]
                        if year == "2024":
                            res["summary"]["republican"] = 8310
                            res["summary"]["democrat"] = 6730
                            res["summary"]["general"] = 350
                        elif year == "2022":
                            res["summary"]["republican"] = 11164
                            res["summary"]["democrat"] = 7928
                            res["summary"]["general"] = 1058
                        elif year == "2018":
                            res["summary"]["republican"] = 11913
                            res["summary"]["democrat"] = 9338
                            res["summary"]["general"] = 107
                            
                    historical[year] = res
                    
    data = {
        "reportTitle": report_title,
        "electionName": election_name,
        "electionDate": election_date,
        "earlyVotingStartDate": early_voting_start,
        "earlyVotingEndDate": early_voting_end,
        "locations": locations,
        "dailyTurnout": daily_turnout,
        "totals": totals,
        "summary": {
            "totalRegistered": total_registered,
            "turnoutPercent": turnout_percent_calc,
            "earlyVoting": early_voting_sum,
            "absenteeNH": absentee_sum,
            "grandTotal": grand_total_calc
        },
        "partyTotals": {
            "republican": republican_total,
            "democrat": democrat_total,
            "general": general_total,
            "total": party_total_sum
        },
        "historical": historical,
        "demographics": demographics,
        "firstTimeVoters": first_time_voters,
        "disclaimer": disclaimer
    }
    
    # Output to data.js
    with open('data.js', 'w', encoding='utf-8') as f:
        f.write("var TURNOUT_DATA = " + json.dumps(data, indent=2) + ";\n")
    print("Successfully generated data.js without hard dependency on Excel sheet")

if __name__ == '__main__':
    main()
